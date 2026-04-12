import json
import os
import re
from io import BytesIO
from uuid import uuid4
from datetime import datetime
from flask import request, jsonify, send_file
from flask_jwt_extended import jwt_required
from admin_api import admin_bp
from extensions import db
from sqlalchemy.orm import joinedload
from models import Customer, ImportHistory
from decorators import admin_required
from utils import paginate_query, apply_filters, apply_sorting, format_paginated, validate_required, validate_alphanumeric, get_or_404, check_unique

IMPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'imports')
ALLOWED_EXCEL_EXTENSIONS = {'xlsx'}
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
EXPORT_COLUMNS = ['customer_id', 'name', 'address', 'created_by_name', 'created_at', 'updated_at']
IMPORT_COLUMNS = ['customer_id', 'name', 'address']


@admin_bp.route('/customers', methods=['GET'])
@jwt_required()
@admin_required
def get_customers(admin):
    query = Customer.query.options(joinedload(Customer.creator))

    filters = {
        'customer_id': {'type': 'fuzzy'},
        'name': {'type': 'fuzzy'},
        'address': {'type': 'fuzzy'},
        'created_at': {'type': 'range', 'cast': lambda x: datetime.fromisoformat(x)},
    }

    query = apply_filters(query, Customer, filters, search_logic='AND')
    query = apply_sorting(query, Customer, sortable_fields=['customer_id', 'name', 'created_at', 'updated_at'], default_sort='-created_at')
    result = paginate_query(query, default_per_page=10)

    return format_paginated('customers', result)


@admin_bp.route('/customers', methods=['POST'])
@jwt_required()
@admin_required
def create_customer(admin):
    if not admin.has_permission('customers', 'create'):
        return jsonify({'message': 'Permission denied'}), 403

    data = request.get_json()

    err = validate_required(data, ['customer_id', 'name'])
    if err: return err

    err = validate_alphanumeric(data['customer_id'], 'Customer ID')
    if err: return err

    err = check_unique(Customer, 'customer_id', data['customer_id'])
    if err: return err

    customer = Customer(
        customer_id=data['customer_id'],
        name=data['name'],
        address=data.get('address', ''),
        created_by=admin.id,
    )

    db.session.add(customer)
    db.session.commit()

    return jsonify(customer.to_dict()), 201


@admin_bp.route('/customers/<customer_public_id>', methods=['GET'])
@jwt_required()
@admin_required
def get_customer(admin, customer_public_id):
    customer, err = get_or_404(Customer, customer_public_id)
    if err: return err
    return jsonify(customer.to_dict()), 200


@admin_bp.route('/customers/<customer_public_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_customer(admin, customer_public_id):
    if not admin.has_permission('customers', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    customer, err = get_or_404(Customer, customer_public_id)
    if err: return err

    data = request.get_json()

    if data.get('customer_id'):
        err = validate_alphanumeric(data['customer_id'], 'Customer ID')
        if err: return err
        err = check_unique(Customer, 'customer_id', data['customer_id'], exclude_id=customer.id)
        if err: return err
        customer.customer_id = data['customer_id']

    if data.get('name'):
        customer.name = data['name']

    if 'address' in data:
        customer.address = data['address']

    db.session.commit()

    return jsonify(customer.to_dict()), 200


@admin_bp.route('/customers/<customer_public_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_customer(admin, customer_public_id):
    if not admin.has_permission('customers', 'delete'):
        return jsonify({'message': 'Permission denied'}), 403

    customer, err = get_or_404(Customer, customer_public_id)
    if err: return err

    db.session.delete(customer)
    db.session.commit()

    return jsonify({'message': 'Customer deleted successfully'}), 200


# ── Export ──

@admin_bp.route('/customers/export', methods=['GET'])
@jwt_required()
@admin_required
def export_customers(admin):
    from openpyxl import Workbook

    customers = Customer.query.options(joinedload(Customer.creator)).order_by(Customer.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Customers'
    ws.append(EXPORT_COLUMNS)

    for c in customers:
        ws.append([
            c.customer_id,
            c.name,
            c.address or '',
            c.creator.name if c.creator else '',
            c.created_at.isoformat() if c.created_at else '',
            c.updated_at.isoformat() if c.updated_at else '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='customers.xlsx')


# ── Import Preview ──

def _parse_excel(file):
    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header:
        return None, 'Excel file is empty'

    header = [str(h).strip().lower() if h else '' for h in header]
    if 'customer_id' not in header or 'name' not in header:
        return None, 'Excel must have columns: customer_id, name'

    col_map = {}
    for col_name in IMPORT_COLUMNS:
        if col_name in header:
            col_map[col_name] = header.index(col_name)

    rows = []
    for i, row in enumerate(rows_iter, start=2):
        data = {}
        for col_name, idx in col_map.items():
            val = row[idx] if idx < len(row) else None
            data[col_name] = str(val).strip() if val is not None else ''
        data['_row'] = i
        rows.append(data)

    wb.close()
    return rows, None


def _validate_import_row(data, existing_map):
    errors = []
    customer_id = data.get('customer_id', '')
    name = data.get('name', '')

    if not customer_id:
        errors.append('customer_id is required')
    elif not re.fullmatch(r'[A-Za-z0-9\-_]+', customer_id):
        errors.append('Customer ID must contain only English letters, numbers, hyphens or underscores')

    if not name:
        errors.append('name is required')

    if errors:
        return 'error', errors, None

    existing = existing_map.get(customer_id)
    if existing:
        existing_name = (existing.name or '').strip()
        existing_address = (existing.address or '').strip()
        import_address = data.get('address', '').strip()
        if existing_name == name and existing_address == import_address:
            return 'unchanged', [], existing
        return 'replace', [], existing
    return 'new', [], None


@admin_bp.route('/customers/import/preview', methods=['POST'])
@jwt_required()
@admin_required
def import_preview(admin):
    if not admin.has_permission('customers', 'create') or not admin.has_permission('customers', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'message': 'No file provided'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        return jsonify({'message': 'Only .xlsx files are allowed'}), 400

    rows, err = _parse_excel(file)
    if err:
        return jsonify({'message': err}), 400

    customer_ids = [r['customer_id'] for r in rows if r.get('customer_id')]
    existing_customers = Customer.query.filter(Customer.customer_id.in_(customer_ids)).all() if customer_ids else []
    existing_map = {c.customer_id: c for c in existing_customers}

    preview_rows = []
    summary = {'total': len(rows), 'new': 0, 'replace': 0, 'unchanged': 0, 'error': 0}

    for data in rows:
        status, errors, existing = _validate_import_row(data, existing_map)
        summary[status] += 1

        row_result = {
            'row': data['_row'],
            'customer_id': data.get('customer_id', ''),
            'name': data.get('name', ''),
            'address': data.get('address', ''),
            'status': status,
            'errors': errors,
        }
        if status == 'replace' and existing:
            row_result['existing'] = {
                'name': existing.name,
                'address': existing.address or '',
            }
        preview_rows.append(row_result)

    return jsonify({'rows': preview_rows, 'summary': summary}), 200


# ── Import Confirm ──

@admin_bp.route('/customers/import/confirm', methods=['POST'])
@jwt_required()
@admin_required
def import_confirm(admin):
    if not admin.has_permission('customers', 'create') or not admin.has_permission('customers', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    file = request.files.get('file')
    data = request.form.get('rows')
    if not data:
        return jsonify({'message': 'No rows provided'}), 400

    try:
        selected_rows = json.loads(data)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'message': 'Invalid rows data'}), 400

    if not selected_rows:
        return jsonify({'message': 'No rows selected'}), 400

    customer_ids = [r['customer_id'] for r in selected_rows]
    existing_customers = Customer.query.filter(Customer.customer_id.in_(customer_ids)).all()
    existing_map = {c.customer_id: c for c in existing_customers}

    created = 0
    updated = 0

    for row in selected_rows:
        existing = existing_map.get(row['customer_id'])
        if existing:
            existing.name = row['name']
            existing.address = row.get('address', '')
            updated += 1
        else:
            customer = Customer(
                customer_id=row['customer_id'],
                name=row['name'],
                address=row.get('address', ''),
                created_by=admin.id,
            )
            db.session.add(customer)
            created += 1

    # Save import history
    stored_filename = None
    if file and file.filename:
        os.makedirs(IMPORT_DIR, exist_ok=True)
        ext = file.filename.rsplit('.', 1)[-1].lower()
        stored_filename = f"{uuid4().hex}.{ext}"
        file.save(os.path.join(IMPORT_DIR, stored_filename))

        history = ImportHistory(
            original_filename=file.filename,
            stored_filename=stored_filename,
            imported_by=admin.id,
        )
        db.session.add(history)

    db.session.commit()

    return jsonify({'created': created, 'updated': updated}), 200


# ── Import History ──

@admin_bp.route('/customers/import/history', methods=['GET'])
@jwt_required()
@admin_required
def get_import_history(admin):
    query = ImportHistory.query.options(joinedload(ImportHistory.importer)).order_by(ImportHistory.created_at.desc())
    result = paginate_query(query, default_per_page=10)
    return format_paginated('histories', result)


@admin_bp.route('/customers/import/history/<history_id>/download', methods=['GET'])
@jwt_required()
@admin_required
def download_import_file(admin, history_id):
    history, err = get_or_404(ImportHistory, history_id)
    if err: return err

    filepath = os.path.join(IMPORT_DIR, history.stored_filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found'}), 404

    return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=history.original_filename)


@admin_bp.route('/customers/import/history/<history_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_import_history(admin, history_id):
    if not admin.has_permission('customers', 'delete'):
        return jsonify({'message': 'Permission denied'}), 403

    history, err = get_or_404(ImportHistory, history_id)
    if err: return err

    filepath = os.path.join(IMPORT_DIR, history.stored_filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    db.session.delete(history)
    db.session.commit()

    return jsonify({'message': 'Import history deleted successfully'}), 200
