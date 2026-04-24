import json
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation
from flask import request, jsonify, send_file, g
from flask_jwt_extended import jwt_required
from openpyxl import load_workbook
from admin_api import admin_bp
from extensions import db
from sqlalchemy.orm import joinedload
from models import Parts, ImportHistory
from decorators import admin_required, company_required
from utils import paginate_query, apply_filters, apply_sorting, format_paginated, load_schema, get_or_404_scoped, check_unique
from schemas import PartsCreateSchema, PartsUpdateSchema, PartsResponseSchema, ImportHistoryResponseSchema
from services.import_export import export_to_excel, save_import_history, get_history_or_404, ALLOWED_EXCEL_EXTENSIONS
from config import IMPORT_DIR


RESOURCE_TYPE = 'parts'
EXPORT_COLUMNS = ['parts_code', 'parts_name', 'unit_price', 'created_by_name', 'created_at', 'updated_at']

# Accept both the business-friendly headers the customer uses and snake_case fallbacks.
HEADER_ALIASES = {
    'part no.':    'parts_code',
    'part no':     'parts_code',
    'part_no':     'parts_code',
    'parts_code':  'parts_code',
    'parts code':  'parts_code',
    'parts name':  'parts_name',
    'parts_name':  'parts_name',
    'unit price':  'unit_price',
    'unit_price':  'unit_price',
    'price':       'unit_price',
}

REQUIRED_COLUMNS = ['parts_code', 'parts_name']
IMPORT_COLUMNS = ['parts_code', 'parts_name', 'unit_price']


def _normalize_header(raw):
    return str(raw).strip().lower() if raw else ''


def _to_decimal(v):
    if v is None or v == '':
        return Decimal('0')
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def parse_parts_excel(file):
    """Parse parts Excel with header aliases. Return (rows, err)."""
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        return None, 'Only .xlsx files are allowed'

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header:
        wb.close()
        return None, 'Excel file is empty'

    # Map each raw header to internal field (or skip unknown)
    col_map = {}
    for idx, raw in enumerate(header):
        key = _normalize_header(raw)
        internal = HEADER_ALIASES.get(key)
        if internal and internal not in col_map:
            col_map[internal] = idx

    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        wb.close()
        expected = ', '.join(['PART No.', 'Parts Name', 'UNIT PRICE'])
        return None, f'Excel must have columns: {expected}'

    rows = []
    for i, row in enumerate(rows_iter, start=2):
        data = {'_row': i}
        for internal, idx in col_map.items():
            val = row[idx] if idx < len(row) else None
            if internal == 'unit_price':
                data[internal] = float(_to_decimal(val))
            else:
                data[internal] = str(val).strip() if val is not None else ''
        rows.append(data)

    wb.close()
    return rows, None


def _validate_parts_row(data, existing_map):
    errors = []
    code = data.get('parts_code', '')
    if not code:
        errors.append('parts_code is required')
    if not data.get('parts_name', ''):
        errors.append('parts_name is required')

    if errors:
        return 'error', errors, None

    ex = existing_map.get(code)
    if ex is None:
        return 'new', [], None

    # Compare: same → unchanged; different → replace
    same = (
        ex.parts_name == data.get('parts_name', '')
        and float(ex.unit_price or 0) == float(data.get('unit_price') or 0)
    )
    return ('unchanged' if same else 'replace'), [], ex


# ── CRUD ──

@admin_bp.route('/parts', methods=['GET'])
@jwt_required()
@admin_required
@company_required
def get_parts(admin):
    query = Parts.query.options(joinedload(Parts.creator)).filter(
        Parts.company_id == g.active_company.id,
        Parts.is_deleted == False,
    )

    filters = {
        'parts_code': {'type': 'fuzzy'},
        'parts_name': {'type': 'fuzzy'},
        'created_at': {'type': 'range', 'cast': lambda x: datetime.fromisoformat(x)},
    }

    query = apply_filters(query, Parts, filters, search_logic='AND')
    query = apply_sorting(query, Parts, sortable_fields=['parts_code', 'parts_name', 'unit_price', 'created_at', 'updated_at'], default_sort='-created_at')
    result = paginate_query(query, default_per_page=10)

    return format_paginated('parts', result, schema=PartsResponseSchema())


@admin_bp.route('/parts', methods=['POST'])
@jwt_required()
@admin_required
@company_required
def create_parts(admin):
    if not admin.has_permission('parts', 'create'):
        return jsonify({'message': 'Permission denied'}), 403

    current_count = Parts.query.filter(
        Parts.company_id == g.active_company.id,
        Parts.is_deleted == False,
    ).count()
    if not admin.check_limit('parts', current_count):
        return jsonify({'message': 'Parts limit reached for your package'}), 403

    data, err = load_schema(PartsCreateSchema)
    if err: return err

    err = check_unique(Parts, 'parts_code', data['parts_code'], company_id=g.active_company.id)
    if err: return err

    part = Parts(
        parts_code=data['parts_code'],
        parts_name=data['parts_name'],
        unit_price=data.get('unit_price', 0),
        company_id=g.active_company.id,
        created_by=admin.id,
    )

    db.session.add(part)
    db.session.commit()

    return jsonify(PartsResponseSchema().dump(part)), 201


@admin_bp.route('/parts/<part_public_id>', methods=['GET'])
@jwt_required()
@admin_required
@company_required
def get_part(admin, part_public_id):
    part, err = get_or_404_scoped(Parts, part_public_id, g.active_company, is_deleted=False)
    if err: return err
    return jsonify(PartsResponseSchema().dump(part)), 200


@admin_bp.route('/parts/<part_public_id>', methods=['PUT'])
@jwt_required()
@admin_required
@company_required
def update_part(admin, part_public_id):
    if not admin.has_permission('parts', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    part, err = get_or_404_scoped(Parts, part_public_id, g.active_company, is_deleted=False)
    if err: return err

    data, err = load_schema(PartsUpdateSchema)
    if err: return err

    if data.get('parts_code'):
        err = check_unique(Parts, 'parts_code', data['parts_code'], exclude_id=part.id, company_id=g.active_company.id)
        if err: return err
        part.parts_code = data['parts_code']

    if data.get('parts_name'):
        part.parts_name = data['parts_name']

    if 'unit_price' in data:
        part.unit_price = data['unit_price']

    db.session.commit()

    return jsonify(PartsResponseSchema().dump(part)), 200


@admin_bp.route('/parts/<part_public_id>', methods=['DELETE'])
@jwt_required()
@admin_required
@company_required
def delete_part(admin, part_public_id):
    if not admin.has_permission('parts', 'delete'):
        return jsonify({'message': 'Permission denied'}), 403

    part, err = get_or_404_scoped(Parts, part_public_id, g.active_company, is_deleted=False)
    if err: return err

    # Soft delete — consumption records that reference this part still resolve
    # via SET NULL on parts_id; snapshot fields remain intact.
    part.is_deleted = True
    db.session.commit()

    return jsonify({'message': 'Part deleted successfully'}), 200


# ── Export ──

@admin_bp.route('/parts/export', methods=['GET'])
@jwt_required()
@admin_required
@company_required
def export_parts(admin):
    items = Parts.query.options(joinedload(Parts.creator)) \
        .filter(Parts.company_id == g.active_company.id, Parts.is_deleted == False) \
        .order_by(Parts.created_at.desc()).all()

    def row_mapper(p):
        return [
            p.parts_code,
            p.parts_name,
            float(p.unit_price) if p.unit_price is not None else 0.0,
            p.creator.name if p.creator else '',
            p.created_at.isoformat() if p.created_at else '',
            p.updated_at.isoformat() if p.updated_at else '',
        ]

    return export_to_excel(items, EXPORT_COLUMNS, row_mapper, 'Parts', 'parts.xlsx')


# ── Import ──

@admin_bp.route('/parts/import/preview', methods=['POST'])
@jwt_required()
@admin_required
@company_required
def import_parts_preview(admin):
    if not admin.has_permission('parts', 'create') or not admin.has_permission('parts', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'message': 'No file provided'}), 400

    rows, err = parse_parts_excel(file)
    if err:
        return jsonify({'message': err}), 400

    codes = [r['parts_code'] for r in rows if r.get('parts_code')]
    existing = Parts.query.filter(
        Parts.parts_code.in_(codes),
        Parts.company_id == g.active_company.id,
        Parts.is_deleted == False,
    ).all() if codes else []
    existing_map = {p.parts_code: p for p in existing}

    preview_rows = []
    summary = {'total': len(rows), 'new': 0, 'replace': 0, 'unchanged': 0, 'error': 0}

    for data in rows:
        status, errors, ex = _validate_parts_row(data, existing_map)
        summary[status] += 1

        row_result = {
            'row': data['_row'],
            'parts_code': data.get('parts_code', ''),
            'parts_name': data.get('parts_name', ''),
            'unit_price': data.get('unit_price', 0),
            'status': status,
            'errors': errors,
        }
        if status == 'replace' and ex:
            row_result['existing'] = {
                'parts_name': ex.parts_name,
                'unit_price': float(ex.unit_price or 0),
            }
        preview_rows.append(row_result)

    return jsonify({'rows': preview_rows, 'summary': summary}), 200


@admin_bp.route('/parts/import/confirm', methods=['POST'])
@jwt_required()
@admin_required
@company_required
def import_parts_confirm(admin):
    if not admin.has_permission('parts', 'create') or not admin.has_permission('parts', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    file = request.files.get('file')
    data = request.form.get('rows')
    if not data:
        return jsonify({'message': 'No rows provided'}), 400

    try:
        selected_rows = json.loads(data)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'message': 'Invalid rows data'}), 400

    if not selected_rows:
        return jsonify({'message': 'No rows selected'}), 400

    codes = [r['parts_code'] for r in selected_rows]
    existing = Parts.query.filter(
        Parts.parts_code.in_(codes),
        Parts.company_id == g.active_company.id,
        Parts.is_deleted == False,
    ).all()
    existing_map = {p.parts_code: p for p in existing}

    new_rows_count = sum(1 for r in selected_rows if r['parts_code'] not in existing_map)
    if new_rows_count > 0:
        current_count = Parts.query.filter(
            Parts.company_id == g.active_company.id,
            Parts.is_deleted == False,
        ).count()
        if not admin.check_limit('parts', current_count, add_count=new_rows_count):
            return jsonify({'message': f'Import would exceed parts limit (current: {current_count}, adding: {new_rows_count})'}), 403

    created = 0
    updated = 0

    for row in selected_rows:
        ex = existing_map.get(row['parts_code'])
        unit_price = _to_decimal(row.get('unit_price'))
        if ex:
            ex.parts_name = row['parts_name']
            ex.unit_price = unit_price
            updated += 1
        else:
            db.session.add(Parts(
                parts_code=row['parts_code'],
                parts_name=row['parts_name'],
                unit_price=unit_price,
                company_id=g.active_company.id,
                created_by=admin.id,
            ))
            created += 1

    save_import_history(file, RESOURCE_TYPE, g.active_company.id, admin.id)
    db.session.commit()

    return jsonify({'created': created, 'updated': updated}), 200


# ── Import History ──

@admin_bp.route('/parts/import/history', methods=['GET'])
@jwt_required()
@admin_required
@company_required
def get_parts_import_history(admin):
    query = ImportHistory.query.options(joinedload(ImportHistory.importer)) \
        .filter(ImportHistory.company_id == g.active_company.id, ImportHistory.resource_type == RESOURCE_TYPE) \
        .order_by(ImportHistory.created_at.desc())
    result = paginate_query(query, default_per_page=10)
    return format_paginated('histories', result, schema=ImportHistoryResponseSchema())


@admin_bp.route('/parts/import/history/<history_id>/download', methods=['GET'])
@jwt_required()
@admin_required
@company_required
def download_parts_import_file(admin, history_id):
    history, err = get_history_or_404(history_id, RESOURCE_TYPE, g.active_company)
    if err: return err

    filepath = os.path.join(IMPORT_DIR, history.stored_filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found'}), 404

    return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=history.original_filename)


@admin_bp.route('/parts/import/history/<history_id>', methods=['DELETE'])
@jwt_required()
@admin_required
@company_required
def delete_parts_import_history(admin, history_id):
    if not admin.has_permission('parts', 'delete'):
        return jsonify({'message': 'Permission denied'}), 403

    history, err = get_history_or_404(history_id, RESOURCE_TYPE, g.active_company)
    if err: return err

    filepath = os.path.join(IMPORT_DIR, history.stored_filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    db.session.delete(history)
    db.session.commit()

    return jsonify({'message': 'Import history deleted successfully'}), 200
