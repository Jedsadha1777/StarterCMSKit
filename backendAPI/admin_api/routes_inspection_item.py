import json
import os
import re
from io import BytesIO
from uuid import uuid4
from datetime import datetime
from flask import request, jsonify, send_file
from flask_jwt_extended import jwt_required
from admin_api import admin_bp
from extensions import db
from sqlalchemy.orm import joinedload
from models import InspectionItem, ImportHistory
from decorators import admin_required
from utils import paginate_query, apply_filters, apply_sorting, format_paginated, validate_required, validate_alphanumeric, get_or_404, check_unique

IMPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'imports')
ALLOWED_EXCEL_EXTENSIONS = {'xlsx'}
EXPORT_COLUMNS = ['item_code', 'item_name', 'spec', 'created_by_name', 'created_at', 'updated_at']
IMPORT_COLUMNS = ['item_code', 'item_name', 'spec']


# ── CRUD ──

@admin_bp.route('/inspection-items', methods=['GET'])
@jwt_required()
@admin_required
def get_inspection_items(admin):
    query = InspectionItem.query.options(joinedload(InspectionItem.creator))

    filters = {
        'item_code': {'type': 'fuzzy'},
        'item_name': {'type': 'fuzzy'},
        'spec': {'type': 'fuzzy'},
        'created_at': {'type': 'range', 'cast': lambda x: datetime.fromisoformat(x)},
    }

    query = apply_filters(query, InspectionItem, filters, search_logic='AND')
    query = apply_sorting(query, InspectionItem, sortable_fields=['item_code', 'item_name', 'created_at', 'updated_at'], default_sort='-created_at')
    result = paginate_query(query, default_per_page=10)

    return format_paginated('inspection_items', result)


@admin_bp.route('/inspection-items', methods=['POST'])
@jwt_required()
@admin_required
def create_inspection_item(admin):
    if not admin.has_permission('inspection_items', 'create'):
        return jsonify({'message': 'Permission denied'}), 403

    data = request.get_json()

    err = validate_required(data, ['item_code', 'item_name'])
    if err: return err

    err = validate_alphanumeric(data['item_code'], 'Item Code')
    if err: return err

    err = check_unique(InspectionItem, 'item_code', data['item_code'])
    if err: return err

    item = InspectionItem(
        item_code=data['item_code'],
        item_name=data['item_name'],
        spec=data.get('spec', ''),
        created_by=admin.id,
    )

    db.session.add(item)
    db.session.commit()

    return jsonify(item.to_dict()), 201


@admin_bp.route('/inspection-items/<item_public_id>', methods=['GET'])
@jwt_required()
@admin_required
def get_inspection_item(admin, item_public_id):
    item, err = get_or_404(InspectionItem, item_public_id)
    if err: return err
    return jsonify(item.to_dict()), 200


@admin_bp.route('/inspection-items/<item_public_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_inspection_item(admin, item_public_id):
    if not admin.has_permission('inspection_items', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    item, err = get_or_404(InspectionItem, item_public_id)
    if err: return err

    data = request.get_json()

    if data.get('item_code'):
        err = validate_alphanumeric(data['item_code'], 'Item Code')
        if err: return err
        err = check_unique(InspectionItem, 'item_code', data['item_code'], exclude_id=item.id)
        if err: return err
        item.item_code = data['item_code']

    if data.get('item_name'):
        item.item_name = data['item_name']

    if 'spec' in data:
        item.spec = data['spec']

    db.session.commit()

    return jsonify(item.to_dict()), 200


@admin_bp.route('/inspection-items/<item_public_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_inspection_item(admin, item_public_id):
    if not admin.has_permission('inspection_items', 'delete'):
        return jsonify({'message': 'Permission denied'}), 403

    item, err = get_or_404(InspectionItem, item_public_id)
    if err: return err

    db.session.delete(item)
    db.session.commit()

    return jsonify({'message': 'Inspection item deleted successfully'}), 200


# ── Export ──

@admin_bp.route('/inspection-items/export', methods=['GET'])
@jwt_required()
@admin_required
def export_inspection_items(admin):
    from openpyxl import Workbook

    items = InspectionItem.query.options(joinedload(InspectionItem.creator)).order_by(InspectionItem.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inspection Items'
    ws.append(EXPORT_COLUMNS)

    for item in items:
        ws.append([
            item.item_code,
            item.item_name,
            item.spec or '',
            item.creator.name if item.creator else '',
            item.created_at.isoformat() if item.created_at else '',
            item.updated_at.isoformat() if item.updated_at else '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='inspection_items.xlsx')


# ── Import Preview ──

def _parse_inspection_excel(file):
    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header:
        return None, 'Excel file is empty'

    header = [str(h).strip().lower() if h else '' for h in header]
    if 'item_code' not in header or 'item_name' not in header:
        return None, 'Excel must have columns: item_code, item_name'

    col_map = {}
    for col_name in IMPORT_COLUMNS:
        if col_name in header:
            col_map[col_name] = header.index(col_name)

    rows = []
    for i, row in enumerate(rows_iter, start=2):
        data = {}
        for col_name, idx in col_map.items():
            val = row[idx] if idx < len(row) else None
            data[col_name] = str(val).strip() if val is not None else ''
        data['_row'] = i
        rows.append(data)

    wb.close()
    return rows, None


def _validate_inspection_row(data, existing_map):
    errors = []
    item_code = data.get('item_code', '')
    item_name = data.get('item_name', '')

    if not item_code:
        errors.append('item_code is required')
    elif not re.fullmatch(r'[A-Za-z0-9\-_]+', item_code):
        errors.append('Item Code must contain only English letters, numbers, hyphens or underscores')

    if not item_name:
        errors.append('item_name is required')

    if errors:
        return 'error', errors, None

    existing = existing_map.get(item_code)
    if existing:
        existing_name = (existing.item_name or '').strip()
        existing_spec = (existing.spec or '').strip()
        import_spec = data.get('spec', '').strip()
        if existing_name == item_name and existing_spec == import_spec:
            return 'unchanged', [], existing
        return 'replace', [], existing
    return 'new', [], None


@admin_bp.route('/inspection-items/import/preview', methods=['POST'])
@jwt_required()
@admin_required
def import_inspection_preview(admin):
    if not admin.has_permission('inspection_items', 'create') or not admin.has_permission('inspection_items', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'message': 'No file provided'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        return jsonify({'message': 'Only .xlsx files are allowed'}), 400

    rows, err = _parse_inspection_excel(file)
    if err:
        return jsonify({'message': err}), 400

    item_codes = [r['item_code'] for r in rows if r.get('item_code')]
    existing_items = InspectionItem.query.filter(InspectionItem.item_code.in_(item_codes)).all() if item_codes else []
    existing_map = {i.item_code: i for i in existing_items}

    preview_rows = []
    summary = {'total': len(rows), 'new': 0, 'replace': 0, 'unchanged': 0, 'error': 0}

    for data in rows:
        status, errors, existing = _validate_inspection_row(data, existing_map)
        summary[status] += 1

        row_result = {
            'row': data['_row'],
            'item_code': data.get('item_code', ''),
            'item_name': data.get('item_name', ''),
            'spec': data.get('spec', ''),
            'status': status,
            'errors': errors,
        }
        if status == 'replace' and existing:
            row_result['existing'] = {
                'item_name': existing.item_name,
                'spec': existing.spec or '',
            }
        preview_rows.append(row_result)

    return jsonify({'rows': preview_rows, 'summary': summary}), 200


# ── Import Confirm ──

@admin_bp.route('/inspection-items/import/confirm', methods=['POST'])
@jwt_required()
@admin_required
def import_inspection_confirm(admin):
    if not admin.has_permission('inspection_items', 'create') or not admin.has_permission('inspection_items', 'edit'):
        return jsonify({'message': 'Permission denied'}), 403

    file = request.files.get('file')
    data = request.form.get('rows')
    if not data:
        return jsonify({'message': 'No rows provided'}), 400

    try:
        selected_rows = json.loads(data)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'message': 'Invalid rows data'}), 400

    if not selected_rows:
        return jsonify({'message': 'No rows selected'}), 400

    item_codes = [r['item_code'] for r in selected_rows]
    existing_items = InspectionItem.query.filter(InspectionItem.item_code.in_(item_codes)).all()
    existing_map = {i.item_code: i for i in existing_items}

    created = 0
    updated = 0

    for row in selected_rows:
        existing = existing_map.get(row['item_code'])
        if existing:
            existing.item_name = row['item_name']
            existing.spec = row.get('spec', '')
            updated += 1
        else:
            item = InspectionItem(
                item_code=row['item_code'],
                item_name=row['item_name'],
                spec=row.get('spec', ''),
                created_by=admin.id,
            )
            db.session.add(item)
            created += 1

    stored_filename = None
    if file and file.filename:
        os.makedirs(IMPORT_DIR, exist_ok=True)
        ext = file.filename.rsplit('.', 1)[-1].lower()
        stored_filename = f"{uuid4().hex}.{ext}"
        file.save(os.path.join(IMPORT_DIR, stored_filename))

        history = ImportHistory(
            original_filename=file.filename,
            stored_filename=stored_filename,
            imported_by=admin.id,
        )
        db.session.add(history)

    db.session.commit()

    return jsonify({'created': created, 'updated': updated}), 200


# ── Import History ──

@admin_bp.route('/inspection-items/import/history', methods=['GET'])
@jwt_required()
@admin_required
def get_inspection_import_history(admin):
    query = ImportHistory.query.options(joinedload(ImportHistory.importer)).order_by(ImportHistory.created_at.desc())
    result = paginate_query(query, default_per_page=10)
    return format_paginated('histories', result)


@admin_bp.route('/inspection-items/import/history/<history_id>/download', methods=['GET'])
@jwt_required()
@admin_required
def download_inspection_import_file(admin, history_id):
    history, err = get_or_404(ImportHistory, history_id)
    if err: return err

    filepath = os.path.join(IMPORT_DIR, history.stored_filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found'}), 404

    return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=history.original_filename)


@admin_bp.route('/inspection-items/import/history/<history_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_inspection_import_history(admin, history_id):
    if not admin.has_permission('inspection_items', 'delete'):
        return jsonify({'message': 'Permission denied'}), 403

    history, err = get_or_404(ImportHistory, history_id)
    if err: return err

    filepath = os.path.join(IMPORT_DIR, history.stored_filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    db.session.delete(history)
    db.session.commit()

    return jsonify({'message': 'Import history deleted successfully'}), 200
