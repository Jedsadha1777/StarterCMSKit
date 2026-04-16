import os
import re
from io import BytesIO
from uuid import uuid4
from flask import send_file
from extensions import db
from models import ImportHistory
from config import IMPORT_DIR, ALLOWED_EXCEL_EXTENSIONS

ALPHANUMERIC_RE = re.compile(r'[A-Za-z0-9\-_]+')


def parse_excel(file, required_columns, import_columns):
    from openpyxl import load_workbook

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        return None, 'Only .xlsx files are allowed'

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header:
        wb.close()
        return None, 'Excel file is empty'

    header = [str(h).strip().lower() if h else '' for h in header]
    missing = [c for c in required_columns if c not in header]
    if missing:
        wb.close()
        return None, f'Excel must have columns: {", ".join(missing)}'

    col_map = {col: header.index(col) for col in import_columns if col in header}

    rows = []
    for i, row in enumerate(rows_iter, start=2):
        data = {}
        for col_name, idx in col_map.items():
            val = row[idx] if idx < len(row) else None
            data[col_name] = str(val).strip() if val is not None else ''
        data['_row'] = i
        rows.append(data)

    wb.close()
    return rows, None


def validate_row(data, existing_map, id_field, required_fields, compare_fields):
    errors = []
    id_value = data.get(id_field, '')

    if not id_value:
        errors.append(f'{id_field} is required')
    elif not ALPHANUMERIC_RE.fullmatch(id_value):
        errors.append(f'{id_field} must contain only English letters, numbers, hyphens or underscores')

    for field in required_fields:
        if field != id_field and not data.get(field, ''):
            errors.append(f'{field} is required')

    if errors:
        return 'error', errors, None

    existing = existing_map.get(id_value)
    if existing:
        changed = False
        for field in compare_fields:
            if field == id_field:
                continue
            existing_val = (getattr(existing, field, '') or '').strip()
            import_val = data.get(field, '').strip()
            if existing_val != import_val:
                changed = True
                break
        return ('replace' if changed else 'unchanged'), [], existing

    return 'new', [], None


def export_to_excel(items, columns, row_mapper, sheet_name, download_name):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)

    for item in items:
        ws.append(row_mapper(item))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=download_name)


def get_history_or_404(history_id, resource_type, active_company):
    from utils import get_or_404_scoped
    from flask import jsonify
    history, err = get_or_404_scoped(ImportHistory, history_id, active_company)
    if err:
        return None, err
    if history.resource_type != resource_type:
        return None, (jsonify({'message': 'Import history not found'}), 404)
    return history, None


def save_import_history(file, resource_type, company_id, admin_id):
    if not file or not file.filename:
        return None

    os.makedirs(IMPORT_DIR, exist_ok=True)
    ext = file.filename.rsplit('.', 1)[-1].lower()
    stored_filename = f"{uuid4().hex}.{ext}"
    file.save(os.path.join(IMPORT_DIR, stored_filename))

    history = ImportHistory(
        resource_type=resource_type,
        original_filename=file.filename,
        stored_filename=stored_filename,
        company_id=company_id,
        imported_by=admin_id,
    )
    db.session.add(history)
    return history
